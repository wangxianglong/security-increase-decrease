import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from multiprocessing import Process

root = tk.Tk()
root.geometry("580x350+50+50")  # widthxheight+x+y
root.title("社保公积金增减员生成器")
root.resizable(False, False)

xinyue_path = tk.StringVar()  # 馨悦-供应链项目花名册
tianan_path = tk.StringVar()  # 天安-供应链项目花名册
zhijian_path = tk.StringVar()  # 智建-供应链项目花名册


def select_file_xinyue():
    selected_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    xinyue_path.set(selected_file_path)
    if len(xinyue_path.get()) > 0 and len(tianan_path.get()) > 0:
        button1.config(state=tk.ACTIVE)
    messagebox.showwarning("馨悦-供应链项目花名册是", f'{selected_file_path}\n\n如果文件不正确可重新选择')


def select_file_tianan():
    selected_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    tianan_path.set(selected_file_path)
    if len(xinyue_path.get()) > 0 and len(tianan_path.get()) > 0:
        button1.config(state=tk.ACTIVE)
    messagebox.showwarning("天安-供应链项目花名册是", f'{selected_file_path}\n\n如果文件不正确可重新选择')


def select_file_zhijian():
    selected_file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    zhijian_path.set(selected_file_path)
    if len(xinyue_path.get()) > 0 and len(tianan_path.get()) > 0:
        button1.config(state=tk.ACTIVE)
    messagebox.showwarning("智建-供应链项目花名册是", f'{selected_file_path}\n\n如果文件不正确可重新选择')


wb_anhui_declaration = None  # 安徽和众社保申报Excel

wb_xinyue_accident_insurance = None  # 馨悦意外险Excel
wb_xinyue_increase_decrease = None  # 馨悦增减员申报Excel

wb_tianan_accident_insurance = None  # 天安意外险Excel
wb_tianan_increase_decrease = None  # 天安增减员申报Excel

wb_zhijian_accident_insurance = None  # 智建意外险Excel
wb_zhijian_increase_decrease = None  # 智建增减员申报Excel


def validate_date(date_string):
    date_pattern = r'\d{4}/\d{1,2}/\d{1,2}'
    if re.match(date_pattern, date_string):
        return True
    else:
        return False


# 生成安徽和众的文件
def generate_excel_anhui():
    selected_year = year_combo.get()
    selected_month = month_combo.get().replace('月', '')
    # print(get_days_of_month(int(selected_year), int(selected_month)))

    increase_count = 0  # 增员人数

    decrease_count = 0  # 减员人数

    wb_roster_xinyue = load_workbook(filename=xinyue_path.get(), read_only=True, data_only=True)  # 读取馨悦-供应链项目花名册Excel
    ####################################在职员工##########################################
    sheet_inservice_xinyue = wb_roster_xinyue["花名册在职模板"]  # 读取在职人员
    for row_index in range(1, sheet_inservice_xinyue.max_row + 1):
        join_date = sheet_inservice_xinyue.cell(row=row_index, column=10).value  # 入司日期
        if validate_date(join_date) is True:
            # 判断是否本年本月
            join_date_array = join_date.split("/")
            if int(selected_year) == int(join_date_array[0]) and int(selected_month) == int(join_date_array[1]):
                # 需要买意外险
                company_name = sheet_inservice_xinyue.cell(row=row_index, column=39).value  # 签订合同主体单位名称

                if company_name == '安徽和众企业服务有限公司':
                    increase_count = increase_count + 1
                    if wb_anhui_accident_insurance is None:
                        template_path = os.path.join(os.path.dirname(os.path.realpath(__file__)),'template_file',
                                                     '安徽和众社保申报表.xlsx')
                        file_name = f"安徽和众：《社保增减员申报表》--供应链项目{company_name}{selected_year}年{selected_month}月.xlsx"
                        wb_anhui_accident_insurance = load_workbook(template_path)
                        wb_anhui_accident_insurance.save(file_name)
                        wb_anhui_accident_insurance.close()
                        wb_anhui_accident_insurance = load_workbook(file_name)
                        

                    sheet_anhui_increase = wb_anhui_accident_insurance["增员"]

                    sheet_anhui_increase.cell(5 + increase_count,1).value = increase_count # 序号
                    sheet_anhui_increase.cell(5 + increase_count,2).value = f"{selected_year}年{selected_month}月" # 增员申报时间



     ####################################在职员工##########################################        
  

# 生成天安的文件
def generate_excel_tianan():
    generate_insurance_fund(tianan_path,"广东天安智慧保安服务有限公司",wb_tianan_accident_insurance,wb_tianan_increase_decrease)


# 生成智建的文件
def generate_excel_zhijian():
    generate_insurance_fund(zhijian_path,"广东智建工程有限公司",wb_zhijian_accident_insurance,wb_zhijian_increase_decrease)


# 生成馨悦的文件
def generate_excel_xinyue():
    generate_insurance_fund(xinyue_path,"广州馨悦商务服务有限公司",wb_xinyue_accident_insurance,wb_xinyue_increase_decrease)


# 生成意外险，社保公积金增减员Excel文件
def generate_insurance_fund(roster_file_path, cur_company_name, wb_accident_insurance, wb_increase_decrease):
    selected_year = year_combo.get()
    selected_month = month_combo.get().replace('月', '')
    # print(get_days_of_month(int(selected_year), int(selected_month)))

    wb_roster = load_workbook(filename=roster_file_path.get(), read_only=True, data_only=True)  # 读取馨悦-供应链项目花名册Excel

    accident_increase_count = 0  # 意外险增员人数
    accident_decrease_count = 0  # 意外险减员人数

    social_count = 0  # 社保增/减员人数

    fund_count = 0  # 公积金增/减员人数

    ########################################在职人员################################################
    sheet_inservice = wb_roster["花名册在职模板"]  # 读取在职人员
    for row_index in range(1, sheet_inservice.max_row + 1):
        join_date = sheet_inservice.cell(row=row_index, column=10).value  # 入司日期
        if validate_date(join_date) is True:
            # 判断是否本年本月
            join_date_array = join_date.split("/")
            if int(selected_year) == int(join_date_array[0]) and int(selected_month) == int(join_date_array[1]):
                # 只要当月入职的就买意外险
                company_name = sheet_inservice.cell(row=row_index, column=39).value  # 签订合同主体单位名称
                if company_name == cur_company_name:
                    accident_increase_count = accident_increase_count + 1
                    if wb_accident_insurance is None:
                        template_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'template_file',
                                                     '意外险.xlsx')
                        file_name = f"{company_name}{selected_year}年{selected_month}月 意外险.xlsx"
                        wb_accident_insurance = load_workbook(template_path)
                        wb_accident_insurance.save(file_name)
                        wb_accident_insurance.close()
                        wb_accident_insurance = load_workbook(file_name)

                    sheet_accident_increase = wb_accident_insurance["员工加保"]

                    sheet_accident_increase.cell(4 + accident_increase_count, 1).value = accident_increase_count  # 序号
                    sheet_accident_increase.cell(4 + accident_increase_count, 3).value = sheet_inservice.cell(
                        row=row_index, column=3).value  # 员工姓名
                    sheet_accident_increase.cell(4 + accident_increase_count, 4).value = sheet_inservice.cell(
                        row=row_index, column=19).value  # 员工证件号码
                    sheet_accident_increase.cell(4 + accident_increase_count, 5).value = "身份证"
                    sheet_accident_increase.cell(4 + accident_increase_count, 6).value = sheet_inservice.cell(
                        row=row_index, column=4).value  # 员工出生日期
                    sheet_accident_increase.cell(4 + accident_increase_count, 7).value = sheet_inservice.cell(
                        row=row_index, column=17).value  # 员工出生日期
                    sheet_accident_increase.cell(4 + accident_increase_count, 9).value = '物业服务'
                    sheet_accident_increase.cell(4 + accident_increase_count, 10).value = join_date  # 入职日期
                    sheet_accident_increase.cell(4 + accident_increase_count, 13).value = company_name  # 成本中心

                    if int(join_date_array[2]) <= 15:  # 本月15日前入职，需要社保增员
                        social_count = write_social_declaration(cur_company_name, join_date, None,"新增",row_index,
                                                 selected_month, selected_year, sheet_inservice,
                                                 social_count, wb_increase_decrease)

            elif int(selected_year) == int(join_date_array[0]) and int(selected_month) - 1 == int(join_date_array[1]) and int(join_date_array[2]) > 15:  # 上个月15日后入职，需要社保增员
                social_count = write_social_declaration(cur_company_name, join_date, None,"新增",row_index,
                                         selected_month, selected_year, sheet_inservice,
                                         social_count, wb_increase_decrease)

        regularization_date = sheet_inservice.cell(row=row_index, column=14).value  # 实际转正时间
        if validate_date(regularization_date) is True:
            # 判断是否本年本月
            regularization_date_array = regularization_date.split("/")
            if int(selected_year) == int(regularization_date_array[0]) and int(selected_month) == int(
                    regularization_date_array[1]):
                if int(regularization_date_array[2]) <= 15:  # 本月15日前转正,公积金增员
                    fund_count = write_fund_declaration(cur_company_name, join_date, regularization_date,None,"新增公积金", row_index,
                                           selected_month, selected_year, sheet_inservice,
                                           fund_count, wb_increase_decrease)

            elif int(selected_year) == int(regularization_date_array[0]) and int(selected_month) - 1 == int(regularization_date_array[1]) and int(regularization_date_array[2]) > 15:  # 上个月15日后转正,公积金增员
                fund_count = write_fund_declaration(cur_company_name, join_date, regularization_date,None,"新增公积金", row_index,
                                       selected_month, selected_year, sheet_inservice,
                                       fund_count, wb_increase_decrease)

    ########################################在职人员################################################

    ########################################离职人员################################################
    sheet_dimission = wb_roster["花名册离职模板"]  # 读取离职人员
    for row_index in range(1, sheet_dimission.max_row + 1):
        resignation_date = sheet_dimission.cell(row=row_index, column=56).value  # 离职时间
        if validate_date(resignation_date) is True:
            # 判断是否本年本月
            resignation_date_array = resignation_date.split("/")
            if int(selected_year) == int(resignation_date_array[0]) and int(selected_month) == int(resignation_date_array[1]):
                # 只要当月离职的就减人意外险
                company_name = sheet_dimission.cell(row=row_index, column=39).value  # 签订合同主体单位名称
                if company_name == cur_company_name:
                    accident_decrease_count = accident_decrease_count + 1
                    if wb_accident_insurance is None:
                        template_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'template_file',
                                                     '意外险.xlsx')
                        file_name = f"{company_name}{selected_year}年{selected_month}月 意外险.xlsx"
                        wb_accident_insurance = load_workbook(template_path)
                        wb_accident_insurance.save(file_name)
                        wb_accident_insurance.close()
                        wb_accident_insurance = load_workbook(file_name)

                    sheet_accident_decrease = wb_accident_insurance["减人"]   

                    sheet_accident_decrease.cell(5 + accident_decrease_count,1).value = accident_decrease_count # 序号
                    sheet_accident_decrease.cell(5 + accident_decrease_count,2).value = sheet_dimission.cell(
                        row=row_index, column=3).value  # 员工姓名
                    sheet_accident_decrease.cell(5 + accident_decrease_count,3).value = sheet_dimission.cell(
                        row=row_index, column=19).value # 员工证件号码
                    sheet_accident_decrease.cell(5 + accident_decrease_count,4).value = sheet_dimission.cell(
                        row=row_index, column=17).value # 员工出生日期
                    sheet_accident_decrease.cell(5 + accident_decrease_count,4).value = resignation_date # 离职日期
    
                if int(resignation_date_array[2]) <= 15:  # 本月15日前离职，需要社保/公积金减员
                    social_count = write_social_declaration(cur_company_name, join_date, resignation_date,"停保",row_index,
                                                 selected_month, selected_year, sheet_inservice,
                                                 social_count, wb_increase_decrease)
                    fund_count = write_fund_declaration(cur_company_name, join_date, regularization_date,resignation_date,"停公积金", row_index,
                                       selected_month, selected_year, sheet_inservice,
                                       fund_count, wb_increase_decrease)

            elif int(selected_year) == int(resignation_date_array[0]) and int(selected_month) - 1 == int(resignation_date_array[1]) and int(resignation_date_array[2]) > 15:  # 上个月15日后离职,社保/公积金减员
                social_count = write_social_declaration(cur_company_name, join_date, resignation_date,"停保",row_index,
                                                 selected_month, selected_year, sheet_inservice,
                                                 social_count, wb_increase_decrease)
                fund_count = write_fund_declaration(cur_company_name, join_date, regularization_date,resignation_date,"停公积金", row_index,
                                       selected_month, selected_year, sheet_inservice,
                                       fund_count, wb_increase_decrease)
    ########################################离职人员################################################

    # messagebox.showinfo("提示", f"生成文件【{file_name}】成功")

# 写入社保申报表sheet页
def write_social_declaration(cur_company_name, join_date,resignation_date,description, row_index, selected_month, selected_year,
                             sheet_inservice, social_count, wb_increase_decrease):
    company_name = sheet_inservice.cell(row=row_index, column=39).value  # 签订合同主体单位名称
    if company_name == cur_company_name:
        social_count = social_count + 1
        if wb_increase_decrease is None:
            template_path = os.path.join(os.path.dirname(os.path.realpath(__file__)),
                                         'template_file', '社保公积金增减员申报表.xlsx')
            file_name = f"{company_name}-{selected_year}年{selected_month}月份社保公积金增减员申报表.xlsx"
            wb_increase_decrease = load_workbook(template_path)
            wb_increase_decrease.save(file_name)
            wb_increase_decrease.close()
            wb_increase_decrease = load_workbook(file_name)

        sheet_social_declaration = wb_increase_decrease["社保申报表"]

        sheet_social_declaration.cell(2 + social_count, 1).value = social_count  # 序号
        sheet_social_declaration.cell(2 + social_count,
                                      2).value = f"{selected_year}年{selected_month}月"  # 起始月份
        sheet_social_declaration.cell(2 + social_count, 3).value = '供应链项目'  # 项目名称
        sheet_social_declaration.cell(2 + social_count, 4).value = '广州'  # 社保缴费地
        sheet_social_declaration.cell(2 + social_count, 5).value = description  # 新增/停保
        if description == "新增":
            sheet_social_declaration.cell(2 + social_count, 5).fill = PatternFill(start_color='FFFF00',
                                                                              fill_type='solid') # 填充单元格颜色
        elif description == "停保":
            sheet_social_declaration.cell(2 + social_count, 5).fill = PatternFill(start_color='FABF8F',
                                                                              fill_type='solid')  # 填充单元格颜色
            sheet_social_declaration.cell(2 + social_count, 13).value = resignation_date  # 离职日期
        sheet_social_declaration.cell(2 + social_count, 6).value = sheet_inservice.cell(
            row=row_index, column=3).value  # 姓名
        sheet_social_declaration.cell(2 + social_count, 7).value = sheet_inservice.cell(
            row=row_index, column=8).value  # 部门
        sheet_social_declaration.cell(2 + social_count, 8).value = sheet_inservice.cell(
            row=row_index, column=9).value  # 岗位
        sheet_social_declaration.cell(2 + social_count, 9).value = sheet_inservice.cell(
            row=row_index, column=19).value  # 身份证号码
        sheet_social_declaration.cell(2 + social_count, 10).value = sheet_inservice.cell(
            row=row_index, column=20).value  # 电话号码
        sheet_social_declaration.cell(2 + social_count, 12).value = join_date  # 入职日期
    wb_increase_decrease.save(file_name)
    wb_increase_decrease.close()
    return social_count


# 写入公积金申报表sheet页
def write_fund_declaration(cur_company_name, join_date, regularization_date,resignation_date,description, row_index, selected_month, selected_year,
                           sheet_inservice, fund_count, wb_increase_decrease):
    company_name = sheet_inservice.cell(row=row_index, column=39).value  # 签订合同主体单位名称
    if company_name == cur_company_name:
        fund_count = fund_count + 1
        if wb_increase_decrease is None:
            template_path = os.path.join(os.path.dirname(os.path.realpath(__file__)),
                                         'template_file', '社保公积金增减员申报表.xlsx')
            file_name = f"{company_name}-{selected_year}年{selected_month}月份社保公积金增减员申报表.xlsx"
            wb_increase_decrease = load_workbook(template_path)
            wb_increase_decrease.save(file_name)
            wb_increase_decrease.close()
            wb_increase_decrease = load_workbook(file_name)

        sheet_fund_declaration = wb_increase_decrease["公积金申报表"]

        sheet_fund_declaration.cell(2 + fund_count, 1).value = fund_count  # 序号
        sheet_fund_declaration.cell(2 + fund_count,
                                    2).value = f"{selected_year}年{selected_month}月"  # 起始月份
        sheet_fund_declaration.cell(2 + fund_count, 3).value = '供应链项目'  # 项目名称
        sheet_fund_declaration.cell(2 + fund_count, 4).value = '广州'  # 社保缴费地
        sheet_fund_declaration.cell(2 + fund_count, 5).value = description  # 新增/停保
        if description == '新增公积金':
            sheet_fund_declaration.cell(2 + fund_count, 5).fill = PatternFill(start_color='FFFF00',
                                                                              fill_type='solid') # 填充单元格颜色
        elif description == '停公积金':                                                                     
            sheet_fund_declaration.cell(2 + fund_count, 5).fill = PatternFill(start_color='FABF8F',
                                                                              fill_type='solid') # 填充单元格颜色
            sheet_fund_declaration.cell(2 + fund_count, 14).value = resignation_date  # 离职日期
        sheet_fund_declaration.cell(2 + fund_count, 6).value = sheet_inservice.cell(
            row=row_index, column=3).value  # 姓名
        sheet_fund_declaration.cell(2 + fund_count, 7).value = sheet_inservice.cell(
            row=row_index, column=8).value  # 部门
        sheet_fund_declaration.cell(2 + fund_count, 8).value = sheet_inservice.cell(
            row=row_index, column=9).value  # 岗位
        sheet_fund_declaration.cell(2 + fund_count, 9).value = sheet_inservice.cell(
            row=row_index, column=19).value  # 身份证号码
        sheet_fund_declaration.cell(2 + fund_count, 10).value = sheet_inservice.cell(
            row=row_index, column=20).value  # 电话号码
        sheet_fund_declaration.cell(2 + fund_count, 12).value = sheet_inservice.cell(
            row=row_index, column=22).value  # 婚否
        sheet_fund_declaration.cell(2 + fund_count, 13).value = join_date  # 入职日期
        sheet_fund_declaration.cell(2 + fund_count, 15).value = regularization_date  # 转正日期
    wb_increase_decrease.save(file_name)
    wb_increase_decrease.close()
    return fund_count

# 生成所有的Excel文件
def generate_all_excel():
  
    process_tianan = Process(target=generate_excel_tianan)
    process_zhijian = Process(target=generate_excel_zhijian)
    process_xinyue = Process(target=generate_excel_xinyue)
    process_anhui = Process(target=generate_excel_anhui)
 
 
    process_tianan.start()
    process_zhijian.start()
    process_xinyue.start()
    process_anhui.start()
 
   
    process_tianan.join()
    process_zhijian.join()
    process_xinyue.join()
    process_anhui.join()


if __name__ == '__main__':
    years = list(range(2024, 2055))
    year_combo = ttk.Combobox(root, values=years, width=5)
    year_combo.current(0)  # 设置默认选择为"一月"
    year_combo.configure(state="readonly")
    year_combo.grid(column=0, row=0)

    months = ["1月", "2月", "3月", "4月", "5月", "6月",
              "7月", "8月", "9月", "10月", "11月", "12月"]
    month_combo = ttk.Combobox(root, values=months, width=5)
    month_combo.current(0)  # 设置默认选择为"一月"
    month_combo.configure(state="readonly")
    month_combo.grid(column=1, row=0)

    entry = tk.Entry(root, textvariable=xinyue_path, width=45)
    entry.grid(column=0, row=1)
    entry.configure(state="readonly")
    tk.Button(root, text="选择馨悦-供应链项目花名册", command=select_file_xinyue).grid(row=1, column=1)

    entry = tk.Entry(root, textvariable=tianan_path, width=45)
    entry.grid(column=0, row=2)
    entry.configure(state="readonly")
    tk.Button(root, text="选择天安-供应链项目花名册", command=select_file_tianan).grid(row=2, column=1)

    entry = tk.Entry(root, textvariable=zhijian_path, width=45)
    entry.grid(column=0, row=3)
    entry.configure(state="readonly")
    tk.Button(root, text="选择智建-供应链项目花名册", command=select_file_zhijian).grid(row=3, column=1)

    button1 = tk.Button(root, text="1、生成社保公积金增减员文件", command=generate_all_excel)
    button1.grid(row=5, column=1, pady=20)  # 使Button在row=1, column=1的位置
    button1.config(state=tk.DISABLED)

    root.mainloop()
